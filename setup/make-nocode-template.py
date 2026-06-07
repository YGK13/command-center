# ============================================================
# make-nocode-template.py
# Generates the No-Code Command Center spreadsheet template:
# teaching/Command-Center-NoCode-Template.xlsx
# Tabs: Start Here, Companies, Pipeline, Goals, Tasks, Feeds, Dashboard
# Includes health colour rules and a live summary tab with formulas.
# ============================================================

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "teaching", "Command-Center-NoCode-Template.xlsx")

# ---- palette (matches the app) -------------------------------
BG      = "0A1A26"
HEAD_BG = "12283A"
ACCENT  = "0EA5E9"
INK     = "E7EEF5"
MUTED   = "9FB3C6"
GREEN   = "1E7D46"
AMBER   = "8A6D1F"
RED     = "7D2A2A"
GREEN_T = "C6F6D5"
AMBER_T = "FAF0C8"
RED_T   = "F8D7DA"

thin = Side(style="thin", color="2A3F52")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=INK, size=11)
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = border
    ws.row_dimensions[row].height = 24

def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def fill_rows(ws, rows, start=2):
    for r, row in enumerate(rows, start=start):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

wb = Workbook()

# ============================================================
# START HERE
# ============================================================
ws = wb.active
ws.title = "Start Here"
ws.sheet_view.showGridLines = False
ws["B2"] = "COMMAND CENTER"
ws["B2"].font = Font(bold=True, size=22, color="0EA5E9")
ws["B3"] = "No-Code template, run your business on one spreadsheet."
ws["B3"].font = Font(size=12, color="3D5068")
guide = [
    ("How to use this template", True),
    ("1. Fill the Companies tab with your projects or service lines. Use an honest health score 0-100.", False),
    ("2. Add your deals on Pipeline, your goals on Goals, your to-dos on Tasks.", False),
    ("3. List the 3-5 news sources you follow on Feeds, with the decision each one informs.", False),
    ("4. Open the Dashboard tab. It totals everything and shows your top tasks automatically.", False),
    ("5. Set a recurring 7:00 AM calendar event: \"Open Command Center, commit to 3.\" That is your engine.", False),
    ("", False),
    ("The health column turns red below 50, amber 50-69, green 70+. One glance reads the whole portfolio.", False),
    ("", False),
    ("When you outgrow this, the columns map one-to-one onto the Full-Code app's data file (scripts/data.mjs).", False),
]
r = 5
for text, head in guide:
    cell = ws.cell(row=r, column=2, value=text)
    cell.font = Font(bold=head, size=13 if head else 11, color=INK if head else MUTED)
    r += 1
widths(ws, [3, 100])

# ============================================================
# COMPANIES
# ============================================================
ws = wb.create_sheet("Companies")
ws.sheet_view.showGridLines = False
headers = ["Company / Project", "Health (0-100)", "Revenue now", "Revenue target",
           "Pipeline value", "Focus action (the one next move)", "Notes"]
ws.append(headers)
style_header(ws, len(headers))
fill_rows(ws, [
    ["Acme Consulting", 60, 0, 12500, 84000, "Send the Series B proposal you keep dodging", ""],
    ["Side Newsletter", 45, 0, 1000, 0, "Wire up the email signup form", ""],
    ["Advisory Practice", 72, 5000, 8000, 36000, "Convert the two warm intros to calls", ""],
])
widths(ws, [26, 14, 14, 14, 14, 42, 30])
# health colour rules on column B, rows 2-200
rng = "B2:B200"
ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["70"], fill=PatternFill("solid", fgColor=GREEN_T)))
ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["50", "69"], fill=PatternFill("solid", fgColor=AMBER_T)))
ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["50"], fill=PatternFill("solid", fgColor=RED_T)))

# ============================================================
# PIPELINE
# ============================================================
ws = wb.create_sheet("Pipeline")
ws.sheet_view.showGridLines = False
headers = ["Company", "Deal / Contact", "Value", "Stage", "Next action", "Due date"]
ws.append(headers)
style_header(ws, len(headers))
fill_rows(ws, [
    ["Acme Consulting", "Series B SaaS pilot", 84000, "Qualified", "Book discovery call", "2026-06-15"],
    ["Advisory Practice", "PE fund advisory", 36000, "Warm", "Send one-pager", "2026-06-20"],
    ["Acme Consulting", "Enterprise audit", 45000, "Cold", "Referral intro", "2026-06-30"],
])
widths(ws, [22, 30, 12, 14, 36, 14])
# stage helper note
ws["H2"] = "Stages: Cold, Warm, Qualified, Proposal, Closed"
ws["H2"].font = Font(italic=True, color="3D5068", size=10)

# ============================================================
# GOALS
# ============================================================
ws = wb.create_sheet("Goals")
ws.sheet_view.showGridLines = False
headers = ["Company", "Goal / Objective", "Key result", "Current", "Target", "% (auto)"]
ws.append(headers)
style_header(ws, len(headers))
goals = [
    ["Acme Consulting", "Land first pilot, $12.5K MRR", "Discovery calls", 2, 6],
    ["Acme Consulting", "Land first pilot, $12.5K MRR", "Proposals sent", 1, 3],
    ["Acme Consulting", "Land first pilot, $12.5K MRR", "MRR ($)", 0, 12500],
]
fill_rows(ws, goals)
# % formula in col F
for i in range(len(goals)):
    row = 2 + i
    f = ws.cell(row=row, column=6, value=f"=IFERROR(MIN(1,D{row}/E{row}),0)")
    f.number_format = "0%"
    f.border = border
widths(ws, [22, 30, 22, 12, 12, 12])

# ============================================================
# TASKS
# ============================================================
ws = wb.create_sheet("Tasks")
ws.sheet_view.showGridLines = False
headers = ["Priority", "Company", "Task", "Category", "Due date", "Done?"]
ws.append(headers)
style_header(ws, len(headers))
fill_rows(ws, [
    ["Critical", "Acme Consulting", "Send the Series B proposal", "Sales", "2026-06-09", ""],
    ["High", "Advisory Practice", "Prep the PE one-pager", "Sales", "2026-06-11", ""],
    ["Medium", "Side Newsletter", "Draft the welcome email", "Content", "2026-06-14", ""],
])
widths(ws, [12, 22, 40, 14, 14, 8])
# priority colour rules
prng = "A2:A200"
ws.conditional_formatting.add(prng, CellIsRule(operator="equal", formula=['"Critical"'], fill=PatternFill("solid", fgColor=RED_T)))
ws.conditional_formatting.add(prng, CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor=AMBER_T)))

# ============================================================
# FEEDS
# ============================================================
ws = wb.create_sheet("Feeds")
ws.sheet_view.showGridLines = False
headers = ["Source / Topic", "Where (URL or app)", "The decision this informs", "Keep? (review monthly)"]
ws.append(headers)
style_header(ws, len(headers))
fill_rows(ws, [
    ["AI in my industry", "Google Alert / news search", "When to reposition my offer", "Keep"],
    ["Competitor moves", "LinkedIn / news", "What to differentiate against", "Keep"],
    ["My market's funding news", "Industry newsletter", "Who just raised and might buy", "Audition"],
])
widths(ws, [26, 30, 40, 18])

# ============================================================
# DASHBOARD (auto summary)
# ============================================================
ws = wb.create_sheet("Dashboard")
ws.sheet_view.showGridLines = False
ws["B2"] = "DASHBOARD"
ws["B2"].font = Font(bold=True, size=20, color="0EA5E9")
ws["B3"] = "Auto-calculated. Pin this tab. Glance every morning."
ws["B3"].font = Font(size=11, color="3D5068")

def kpi(cell_label, label, cell_val, formula, fmt=None):
    lc = ws[cell_label]; lc.value = label
    lc.font = Font(size=10, bold=True, color=MUTED)
    vc = ws[cell_val]; vc.value = formula
    vc.font = Font(size=18, bold=True, color=INK)
    if fmt: vc.number_format = fmt

kpi("B5", "TOTAL PIPELINE", "B6", "=SUM(Pipeline!C2:C200)", "$#,##0")
kpi("D5", "ACTIVE REVENUE", "D6", "=SUM(Companies!C2:C200)", "$#,##0")
kpi("F5", "OPEN CRITICAL TASKS", "F6", '=COUNTIFS(Tasks!A2:A200,"Critical",Tasks!F2:F200,"")')
kpi("B8", "COMPANIES TRACKED", "B9", "=COUNTA(Companies!A2:A200)")
kpi("D8", "DEALS IN PIPELINE", "D9", "=COUNTA(Pipeline!A2:A200)")
kpi("F8", "AVG HEALTH", "F9", "=IFERROR(ROUND(AVERAGE(Companies!B2:B200),0),0)")

ws["B11"] = "PIPELINE BY COMPANY"
ws["B11"].font = Font(bold=True, size=11, color=ACCENT)
ws["B12"] = "Company"; ws["C12"] = "Pipeline value"
for c in ("B12", "C12"):
    ws[c].font = Font(bold=True, color=MUTED, size=10)
# SUMIF per company row (mirror Companies order)
for i in range(3):
    row = 13 + i
    crow = 2 + i
    ws.cell(row=row, column=2, value=f"=IF(Companies!A{crow}=\"\",\"\",Companies!A{crow})")
    cell = ws.cell(row=row, column=3, value=f"=IF(Companies!A{crow}=\"\",\"\",SUMIF(Pipeline!A:A,Companies!A{crow},Pipeline!C:C))")
    cell.number_format = "$#,##0"

ws["B18"] = "TIP: change the rows in the source tabs and this dashboard updates itself."
ws["B18"].font = Font(italic=True, color="3D5068", size=10)
widths(ws, [3, 22, 18, 4, 20, 18, 22])

wb.save(OUT)
print("Wrote", OUT)
